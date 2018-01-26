#!/usr/bin/python3
import os
import sys
import re
import pandas as pd
import numpy as np
from pandas import ExcelWriter
from optparse import OptionParser
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from collections import Counter
from collections import OrderedDict

defaulturl = '''http://www.moneycontrol.com/india/stockpricequote/miscellaneous/crisil/CRI'''
inxlsfile = os.path.join(os.getcwd(),defaulturl.split()[0].split('/')[6]+"MC.xlsx")

keyreferences = [[],
                 ['EQUITIES AND LIABILITIES',"SHAREHOLDER'S FUNDS",'Total Share Capital',
                  'Total Reserves and Surplus', 'NON-CURRENT LIABILITIES',
                  'Total Non-Current Liabilities', 'CURRENT LIABILITIES', 'Total Current Liabilities',
                  'Total Capital And Liabilities', 'ASSETS', 'NON-CURRENT ASSETS',
                  'Fixed Assets', 'Total Non-Current Assets', 'CURRENT ASSETS',
                  'Total Current Assets', 'Total Assets',
                  'CONTINGENT LIABILITIES, COMMITMENTS', 'CIF VALUE OF IMPORTS',
                  'EXPENDITURE IN FOREIGN EXCHANGE', 'REMITTANCES IN FOREIGN CURRENCIES FOR DIVIDENDS',
                  'EARNINGS IN FOREIGN EXCHANGE', 'BONUS DETAILS',
                  'NON-CURRENT INVESTMENTS', 'CURRENT INVESTMENTS','Source :  Dion Global Solutions Limited'],
                 ['INCOME','Revenue From Operations [Gross]','Revenue From Operations [Net]',
                  'Total Operating Revenues','EXPENSES','Total Expenses',
                  'Profit/Loss Before Tax','Total Tax Expenses','Profit/Loss For The Period',
                  'EARNINGS PER SHARE','VALUE OF IMPORTED AND INDIGENIOUS RAW MATERIALS',
                  'DIVIDEND AND DIVIDEND PERCENTAGE','Equity Dividend Rate (%)','Source :  Dion Global Solutions Limited'],
                 [],[],[],[],
                 ['Net Profit/Loss Before Extraordinary Items And Tax', 'Net Inc/Dec In Cash And Cash Equivalents',
                  'Source :  Dion Global Solutions Limited'],
                 ['Per Share Ratios', 'Profitability Ratios', 
                  'Liquidity Ratios', 'Valuation Ratios','Source :  Dion Global Solutions Limited']
                ]

key_map = {
           'Cash and Equivalents':'Cash And Cash Equivalents',
           'Accounts Receivable, Net':'Trade Receivables',
           'Inventory':'Inventories',
           'Total Current Assets':'Total Current Assets',
           'Net PP&E':'Tangible Assets',
           'Intangible Assets': 'Intangible Assets',
           'Total Assets': 'Total Assets',
           'Accounts Payable':'Trade Payables',
           'Taxes Payable':'Deferred Tax Liabilities [Net]',
           'Total Current Liabilities':'Total Current Liabilities',
           'Long-term Debt':'Long Term Borrowings',
           "Total Stockholder's Equity":'Total Shareholders Funds',
           'Total Liabilities and Equity':'Total Capital And Liabilities',
           'Sales':'Total Operating Revenues',
           'Depreciation and Amortization':'Depreciation And Amortisation Expenses',
           'Interest Expense':'Finance Costs',
           'Other Gains and Losses':'Exceptional Items',
           'Pretax Income': 'Profit/Loss Before Tax',
           'Income Tax Expense':'Total Tax Expenses',
           'Net Income':'Profit/Loss For The Period',
           'Net Cash from Operations':'Net CashFlow From Operating Activities',
           'Net Cash from Investing Activities':'Net Cash Used In Investing Activities',
           'Net Cash from Financing Activities':'Net Cash Used From Financing Activities',
           'Change in cash':'Net Inc/Dec In Cash And Cash Equivalents',
           'Earnings per share': 'Diluted EPS (Rs.)',
           'Dividends per share': 'Dividend / Share(Rs.)',
           'BookValue per share': 'Book Value [InclRevalReserve]/Share (Rs.)',
          }
calc_keymap = {
           'Other Current Assets':[['Total Current Assets'],['Inventories','Trade Receivables','Cash And Cash Equivalents']],
           'Other Current Liabilities':[['Total Current Liabilities'],['Trade Payables']],
           'Other Liabilities': [['Total Non-Current Liabilities'], ['Long Term Borrowings','Deferred Tax Liabilities [Net]']],
           'Total Liabilities': [['Total Current Liabilities','Total Non-Current Liabilities'],[]],
           'Cost of Goods Sold':[['Cost Of Materials Consumed', 
                                  'Purchase Of Stock-In Trade', 
                                  'Changes In Inventories Of FG,WIP And Stock-In Trade'],[]],
           'Gross Profit':[['Total Operating Revenues'],['Cost Of Materials Consumed', 
                            'Purchase Of Stock-In Trade', 'Changes In Inventories Of FG,WIP And Stock-In Trade']],
           'Operating Income before Depr':[['Total Operating Revenues'], ['Cost Of Materials Consumed', 
                             'Purchase Of Stock-In Trade', 'Changes In Inventories Of FG,WIP And Stock-In Trade', 
                             'Employee Benefit Expenses', 'Other Expenses']],
           'Operating Profit':[['Total Operating Revenues'], ['Cost Of Materials Consumed',
                                'Purchase Of Stock-In Trade','Changes In Inventories Of FG,WIP And Stock-In Trade', 
                                'Employee Benefit Expenses', 'Other Expenses', 'Depreciation And Amortisation Expenses']],
           'Selling, General, and Admin Exp':[['Employee Benefit Expenses', 'Other Expenses'],[]]
           }

def read_template(tmpltxlsx,start, end):
    if not os.path.isfile(os.path.join(os.getcwd(),tmpltxlsx)):
        print("Template File Missing")
        return

    df = pd.read_excel(os.path.join(os.getcwd(),tmpltxlsx),index_col=0)
    df.drop(df.columns[:1], axis=1,inplace=True) # Drop the first columns 'Unamed'
    df.index = df.index.str.strip()  # Remove any whitespaces in the index
    df = df.loc[df.index.notnull()]  # Drop all rows with index NaN
    return df[start:end]

def add_data_rows(rowlist, bs, ele):
    if not rowlist: return
    [rowlist.remove(ele) for ele in rowlist[:] if not ele in bs.index]
    df = bs.loc[rowlist,:]
    return df.apply(lambda col: col.sum())

def template_fill(tmpl, bs):
    for ele in tmpl.index:
        if ele in key_map and key_map[ele] in bs.index:
            # if no calculation required, simply copy the values to template
            kmap = key_map[ele]
            tmpl.loc[ele] = bs.loc[kmap].values
            tmpl.loc[ele].fillna('0.00',inplace=True)
        elif ele in calc_keymap:
            # calculate the values before copying to template
            key = calc_keymap[ele]
            bs.loc['scratch1'] = add_data_rows(key[0],bs,ele)
            bs.loc['scratch2'] = add_data_rows(key[1],bs,ele)
            bs.loc['scratch'] = bs.apply(lambda row: row['scratch1'] - row['scratch2'])
            tmpl.loc[ele] = bs.loc['scratch'].values
            tmpl.loc[ele].fillna('0.00',inplace=True)
            bs.drop(['scratch1','scratch2','scratch'],inplace=True)
    return tmpl

def read_xls(xls_path, num_sheets=2):
    df = pd.DataFrame()
    xl = pd.ExcelFile(xls_path)
    for sheetname in xl.sheet_names[:num_sheets]:
        df = df.append(pd.read_excel(xls_path, sheetname=sheetname))
    return df


def save_xls(list_dfs, xls_path):
    writer = ExcelWriter(xls_path)
    for n, df in enumerate(list_dfs):
        df.to_excel(writer,'sheet%s' % n)
    writer.save()

def index_union(df1, df2, pagenum):
    b1list = df1.index.values.tolist()
    b2list = df2.index.values.tolist()
    p1=p2=c1=c2=0
    resultlist = []
    #for ele in (df1.index[df1.isnull().all(1)]): #find all rows with NaN - essentially sub-headers
    for ele in keyreferences[pagenum]:
         if ele in b1list and ele in b2list:
             c1 = b1list.index(ele) 
             c2 = b2list.index(ele) 
             # b1list[p1:c1] - Get all rows between two sub-headers
             # set() ensures unique values and append in the resultlist
             # [1:] ensures skip the sub-header before set() to ensure it doesnt get re-shuffled 
             [resultlist.append(x) for x in b1list[p1:c1][:1]+list(set(b1list[p1:c1][1:]+b2list[p2:c2][1:]))]
             p1 = c1
             p2 = c2
    resultlist = list(OrderedDict.fromkeys(resultlist)) # Remove duplicates
    #print([k for k,v in Counter(resultlist).items() if v>1]) #Find duplicates in resultlist
    return resultlist

def cleanz_data(df):
    df.drop_duplicates(inplace=True)       #Drop duplicate lines, in P&L Year row is duplicated 
    df.fillna(value=np.nan, inplace=True)  #Replace None --> NaN
    df.dropna(how='all', inplace=True)     #Drop rows with all NaN, some tables have NaN row before Year row
    df.iloc[1] = df.iloc[1].shift(1)       #Shift row 2 by 1 column [row 2 is year row]
    df.iat[1,0] = df.iat[0,0]              #Copy title into row 2 before we drop row 1
    df.drop(0, inplace=True)               #Drop row 1 [Redundand row, title is already copied]


def merge_lists(First5yr, Second5yr, pagenum):
    df1 = pd.DataFrame(First5yr); 
    df2 = pd.DataFrame(Second5yr); 
    # Cleanz the data :--> 
    #        No duplicates, 
    #        First row as Year row with title, 
    cleanz_data(df1)
    cleanz_data(df2)
    b1 = df1.dropna(how='all').rename(columns={col : df1.iloc[0][idx] 
                    for idx, col in enumerate(df1.columns) if idx}).set_index(
                    keys=df1.columns[0]) #Make first column as index
    b2 = df2.dropna(how='all').rename(columns={col : df2.iloc[0][idx] 
                    for idx, col in enumerate(df2.columns) if idx}).set_index(
                    keys=df2.columns[0]) #Make first column as index
    # No Year index in any other row -> happens in P&L
    if b1.columns[0] in b1.index : b1.drop(b1.columns[0], inplace=True)
    if b2.columns[0] in b2.index : b2.drop(b2.columns[0], inplace=True)
    return b1.join(b2, how='outer').reindex(index_union(b1, b2, pagenum))

def save_data(data,dest):   
          wb = Workbook()
          ws = wb.active
          for i,row in enumerate(data):
              for j,col in enumerate(row):
                  ws.cell(row=i+1, column=j+1, value=col)
          wb.save(dest)

def pull_data(page_source):
    data = []
    soup = BeautifulSoup(page_source,"lxml")
    divTags = soup.find_all('div', {'class':'boxBg'})
    for tag in divTags:
        tables = tag.find_all('table',attrs={'class':'table4'})
        for table in tables:
            table_bodys = table.find_all('tbody')
            for tablebody in table_bodys:
                rows = tablebody.find_all('tr')
                for row in rows:
                    cols = row.find_all('td')
                    cols = [ele.text.strip() for ele in cols]
                    data.append([ele for ele in cols if ele]) # Get rid of empty values
    return data

def scrape_page(driver,pagenum,cflag):
    navigXpath = ['''//*[@id="slider"]/dt[7]/a''',
                  '''//*[@id="slider"]/dd[3]/ul/li['''+str(pagenum)+''']/a''',
                  '''//*[@id="mc_mainWrapper"]/div[3]/div[2]/div[3]/div[2]/div[2]/div[2]/div[1]/div/ul/li[2]/a''',
		  '''//*[@id="mc_mainWrapper"]/div[3]/div[2]/div[3]/div[2]/div[2]/div[2]/div[1]/div/div[1]/div[1]/table/tbody/tr/td/a/b''']
    try:
        driver.find_element_by_xpath(navigXpath[0]).click()
        driver.find_element_by_xpath(navigXpath[1]).click()
        if not cflag:
            First5yr = pull_data(driver.page_source)
            driver.find_element_by_xpath(navigXpath[3]).click()
            Second5yr = pull_data(driver.page_source)
        else:
            driver.find_element_by_xpath(navigXpath[2]).click()
            First5yr = pull_data(driver.page_source)
            driver.find_element_by_xpath(navigXpath[3]).click()
            Second5yr = pull_data(driver.page_source)
        #save_data(First5yr,defaulturl.split(sep='/')[6] + '1.xlsx')
        #save_data(Second5yr,defaulturl.split(sep='/')[6] + '2.xlsx')
        return merge_lists(First5yr, Second5yr, pagenum)
    except NoSuchElementException:
        print ("Webpage Not Accessible, Try again after some time")
        sys.exit(-1)

def main(options):
    """
    main() function to start the program 
    
    Parameters
    ---------
    url : url to parse  
    
    Returns
    -------
    None 
    """
    tmplrange = ['Balance Sheet', 'Income Statement',
                 'Statement of Cash Flows', 'Shares Outstanding']
    # List of page numbers on moneycontrol website
    if not os.path.isfile(options.inputxlsx):
        pagelist = [1,2,7,8]
        #pagelist = [2]
        driver = webdriver.PhantomJS()
        print("Parsing URL => "+options.url)
        driver.get(options.url)
        listdfs = []
        for pagenum in pagelist:
            listdfs.append(scrape_page(driver,pagenum,options.consoflag))
        save_xls(listdfs, options.inputxlsx)
    bs = read_xls(options.inputxlsx)
    # convert all numbers to float
    for col in bs.columns:
        bs[col] = pd.to_numeric(bs[col],errors='coerce')

    df = pd.DataFrame()
    for idx in range(len(tmplrange)-1):
        template = read_template(options.tmpltxlsx,tmplrange[idx],tmplrange[idx+1])
        df = pd.concat([df,template_fill(template,bs)])
    save_xls([df], options.outtmpltxlsx)
   
def argparser():
    """
    Option parsing of command line

    It will add the required arguments to OptionParser module
    Collects and parse the arguments
    
    Parameters
    ---------
    None 
    
    Returns
    -------
    opts: Parsed arguments (or their defaults) returned in opts
    """
    parser = OptionParser(usage="usage: %prog [options]")
    parser.add_option(
        "-u","--url", dest="url", help="Moneycontrol URL for stock", default=defaulturl)
    parser.add_option(
        "-t","--template", dest="tmpltxlsx", help="Template xlsx file", default='Stock_Analysis_Quant_Template.xlsx')
    parser.add_option(
        "-o","--outtemplate", dest="outtmpltxlsx", help="Out Template xlsx file", default='Out_Quant_Template.xlsx')
    parser.add_option(
        "-x","--inputxls", dest="inputxlsx", help="Input xlsx file", default=inxlsfile)
    parser.add_option(
        "-c","--consolidateflag", action="store_true", dest="consoflag", help="Get consolidated results", default=False)
    (opts, args) = parser.parse_args(sys.argv)
    return opts

if __name__ == '__main__':
     # Parse the arguments and pass to main
     options = argparser()
     sys.exit(main(options))

