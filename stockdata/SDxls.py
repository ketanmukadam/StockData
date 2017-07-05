from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from copy import copy
import re

key_map = {
           'Cash and Equivalents':'Cash And Cash Equivalents',
           'Accounts Receivable, Net':'Trade Receivables',
           'Inventory':'Inventories',
           'Total Current Assets':'Total Current Assets',
           'Net PP&E':'Tangible Assets',
           'Intangible Assets': 'Intangible Assets',
           'Total Assets': 'Total Assets',
           'Accounts Payable':'Trade Payables',
           'Taxes Payable':'Deferred Tax Liabilities [Net]',
           'Total Current Liabilities':'Total Current Liabilities',
           'Long-term Debt':'Long Term Borrowings',
           "Total Stockholder's Equity":'Total Shareholders Funds',
           'Total Liabilities and Equity':'Total Capital And Liabilities',
           'Sales':'Total Operating Revenues',
           'Depreciation and Amortization':'Depreciation And Amortisation Expenses',
           'Interest Expense':'Finance Costs',
           'Other Gains and Losses':'Exceptional Items',
           'Pretax Income': 'Profit/Loss Before Tax',
           'Income Tax Expense':'Total Tax Expenses',
           'Net Income':'Profit/Loss For The Period',
           'Net Cash from Operations':'Net CashFlow From Operating Activities',
           'Net Cash from Investing Activities':'Net Cash Used In Investing Activities',
           'Net Cash from Financing Activities':'Net Cash Used From Financing Activities',
           'Change in cash':'Net Inc/Dec In Cash And Cash Equivalents',
           'Earnings per share': 'Diluted EPS (Rs.)',
           'Dividends per share': 'Dividend / Share(Rs.)',
           'BookValue per share': 'Book Value [InclRevalReserve]/Share (Rs.)',
           'Other Current Assets':'Total Current Assets - Inventories - Trade Receivables - Cash And Cash Equivalents',
           'Other Current Liabilities':'Total Current Liabilities - Trade Payables',
           'Other Liabilities': 'Total Non-Current Liabilities - Long Term Borrowings - Deferred Tax Liabilities [Net]',
           'Total Liabilities': 'Total Current Liabilities + Total Non-Current Liabilities',
           'Cost of Goods Sold':'Cost Of Materials Consumed + Purchase Of Stock-In Trade + Changes In Inventories Of FG,WIP And Stock-In Trade',
           'Gross Profit':'Total Operating Revenues - Cost Of Materials Consumed - Purchase Of Stock-In Trade - Changes In Inventories Of FG,WIP And Stock-In Trade',
           'Operating Income before Depr':'Total Operating Revenues - Cost Of Materials Consumed - Purchase Of Stock-In Trade - Changes In Inventories Of FG,WIP And Stock-In Trade - Employee Benefit Expenses - Other Expenses',
           'Operating Profit':'Total Operating Revenues - Cost Of Materials Consumed - Purchase Of Stock-In Trade - Changes In Inventories Of FG,WIP And Stock-In Trade - Employee Benefit Expenses - Other Expenses - Depreciation And Amortisation Expenses',
           'Selling, General, and Admin Exp':'Employee Benefit Expenses + Other Expenses'
          }

class SDxlsMixin():
      def copy_fulldata(self, dest):
          wb = Workbook()
          ws = wb.active
          for i,row in enumerate(self.data):
              for j,col in enumerate(row):
                  ws.cell(row=i+1, column=j+1, value=col)
          wb.save(dest)

      def oper_list_of_list(self, data, name, oper):
          if oper :
             tempdata = [sum(i) for i in zip(*data) if not str in [type(e) for e in i]] 
          else:
             tempdata = [i[0] - i[1] for i in zip(*data) if not str in [type(e) for e in i]] 
          tempdata.insert(0,name)
          return tempdata
          

      def calculate_datarow(self,key,name):
          datarow = []
          if len(re.split(' \+ | - ', key)) < 2 :
             for drow in self.data:
                  if drow[0] == key:
                      datarow = drow
          else:
             tempdata = []
             delimt = ' - '
             if ' + ' in key : delimt = ' + ' 
             keys = key.split(delimt)
             for k in keys:
                for drow in self.data:
                  if drow[0] == k:
                     tempdata.append(drow)
             if delimt == ' + ':
                tempdata[0] = self.oper_list_of_list(tempdata, name, True)
             if delimt == ' - ':
                tempdata[1] = self.oper_list_of_list(tempdata[1:], name, True)
                tempdata = tempdata[:2]
                tempdata[0] = self.oper_list_of_list(tempdata, name, False) 
             datarow = tempdata[0]
          return datarow

      def copy_cellformat(self,incell, outcell):
          if incell.has_style:
             outcell.font = copy(incell.font)
             outcell.border = copy(incell.border)
             outcell.fill = copy(incell.fill)
             outcell.number_format = copy(incell.number_format)
             outcell.protection = copy(incell.protection)
             outcell.alignment = copy(incell.alignment)
              
      def update_mysheet(self,wb):
          ws = wb.active
          for row in ws.rows:
              if not isinstance(row[0].value,str):continue
              key = key_map.get(row[0].value.strip())
              if not key: continue
              datarow = self.calculate_datarow(key, row[0].value)
              for idx, datacol in enumerate(datarow):
                  if not idx: continue
                  cell = row[idx+1]
                  col = column_index_from_string(cell.column)
                  if type(datacol) != float:
                     newcell = ws.cell(row=cell.row,column=col, value=float(datacol.replace(',','')))
                  else :
                     newcell = ws.cell(row=cell.row,column=col, value=float(datacol))
                     self.copy_cellformat(cell, newcell)
      
      def zap_mysheet(self, ws):
          for row in ws.rows:
              for cell in row:
                  if isinstance(cell.value,float):
                      dcell = ws.cell(row=cell.row, column=column_index_from_string(cell.column), value=0.0)
                      self.copy_cellformat(cell, dcell)

      def copy_mysheet(self,src, dest, sheetname):
          dwb = Workbook()
          dws = dwb.active
          swb = load_workbook(filename = src, keep_vba=True)
          sws = swb.get_sheet_by_name(sheetname)
          dws.title = sws.title
          dws.sheet_view.showGridLines = False
          for row in sws.rows:
              for cell in row:
                  dcell = dws.cell(row=cell.row, column=column_index_from_string(cell.column), value=cell.value)
                  self.copy_cellformat(cell, dcell)
          self.zap_mysheet(dws)
          self.update_mysheet(dwb)
          dwb.save(dest)
