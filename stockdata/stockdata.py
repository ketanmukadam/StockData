import pdb
import csv
import re
import pickle
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException 
from itertools import zip_longest
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from copy import copy


bs_keyref = ['EQUITIES AND LIABILITIES',"SHAREHOLDER'S FUNDS",'Total Share Capital',
    'Total Reserves and Surplus', 'NON-CURRENT LIABILITIES',
    'Total Non-Current Liabilities', 'Total Current Liabilities',
    'Total Capital And Liabilities', 'NON-CURRENT ASSETS',
    'Fixed Assets', 'Total Non-Current Assets',
    'Total Current Assets', 'Total Assets',
    'CONTINGENT LIABILITIES, COMMITMENTS', 'CIF VALUE OF IMPORTS',
    'EXPENDITURE IN FOREIGN EXCHANGE', 'REMITTANCES IN FOREIGN CURRENCIES FOR DIVIDENDS',
    'EARNINGS IN FOREIGN EXCHANGE', 'BONUS DETAILS',
    'NON-CURRENT INVESTMENTS', 'CURRENT INVESTMENTS','Source :  Dion Global Solutions Limited']

pl_keyref = ['Revenue From Operations [Gross]','Revenue From Operations [Net]',
             'Total Operating Revenues','EXPENSES','Total Expenses',
             'Profit/Loss Before Tax','Total Tax Expenses','Profit/Loss For The Period',
             'EARNINGS PER SHARE','VALUE OF IMPORTED AND INDIGENIOUS RAW MATERIALS',
             'DIVIDEND AND DIVIDEND PERCENTAGE','Equity Dividend Rate (%)','Source :  Dion Global Solutions Limited']

cf_keyref = ['Net Profit/Loss Before Extraordinary Items And Tax', 
             'Net Inc/Dec In Cash And Cash Equivalents',
             'Source :  Dion Global Solutions Limited']

ratio_keyref = ['Per Share Ratios', 'Profitability Ratios', 
                'Liquidity Ratios', 'Valuation Ratios','Source :  Dion Global Solutions Limited']

key_map = {
           'Cash and Equivalents':'Cash And Cash Equivalents',
           'Accounts Receivable, Net':'Trade Receivables',
           'Inventory':'Inventories',
           'Total Current Assets':'Total Current Assets',
           'Net PP&E':'Fixed Assets',
           'Intangible Assets': 'Intangible Assets',
           'Total Assets': 'Total Assets',
           'Accounts Payable':'Trade Payables',
           'Taxes Payable':'Deferred Tax Liabilities [Net]',
           'Total Current Liabilities':'Total Current Liabilities',
           'Long-term Debt':'Long Term Borrowings',
           "Total Stockholder's Equity":'Total Shareholders Funds',
           'Total Liabilities and Equity':'Total Capital And Liabilities',
           'Sales':'Total Operating Revenues',
           'Depreciation and Amortization':'Depreciation And Amortisation Expenses',
           'Interest Expense':'Finance Costs',
           'Other Gains and Losses':'Exceptional Items',
           'Pretax Income': 'Profit/Loss Before Tax',
           'Income Tax Expense':'Total Tax Expenses',
           'Net Income':'Profit/Loss For The Period',
           'Net Cash from Operations':'Net CashFlow From Operating Activities',
           'Net Cash from Investing Activities':'Net Cash Used In Investing Activities',
           'Net Cash from Financing Activities':'Net Cash Used From Financing Activities',
           'Change in cash':'Net Inc/Dec In Cash And Cash Equivalents',
           'Earnings per share': 'Diluted EPS (Rs.)',
           'Dividends per share': 'Dividend / Share(Rs.)',
           'BookValue per share': 'Book Value [InclRevalReserve]/Share (Rs.)',
           'Other Current Assets':'Total Current Assets - Inventories - Trade Receivables - Cash And Cash Equivalents',
           'Other Current Liabilities':'Total Current Liabilities - Trade Payables',
           'Other Liabilities': 'Total Non-Current Liabilities - Long Term Borrowings - Deferred Tax Liabilities [Net]',
           'Total Liabilities': 'Total Current Liabilities + Total Non-Current Liabilities',
           'Cost of Goods Sold':'Cost Of Materials Consumed + Purchase Of Stock-In Trade + Changes In Inventories Of FG,WIP And Stock-In Trade',
           'Gross Profit':'Total Operating Revenues - Cost Of Materials Consumed - Purchase Of Stock-In Trade - Changes In Inventories Of FG,WIP And Stock-In Trade',
           'Operating Income before Depr':'Total Operating Revenues - Cost Of Materials Consumed - Purchase Of Stock-In Trade - Changes In Inventories Of FG,WIP And Stock-In Trade - Employee Benefit Expenses - Other Expenses',
           'Operating Profit':'Total Operating Revenues - Cost Of Materials Consumed - Purchase Of Stock-In Trade - Changes In Inventories Of FG,WIP And Stock-In Trade - Employee Benefit Expenses - Other Expenses - Depreciation And Amortisation Expenses',
           'Selling, General, and Admin Exp':'Employee Benefit Expenses + Other Expenses'
          }

class WgetFinData():

      def __init__(self, url, title):
          self.data = []
          self.url = url
          self.start_driver()
          assert title in self.driver.title

      def dump_data(self, filename):
          pickle.dump(self.data, open(filename, "wb"))

      def get_data(self, filename):
          page_list = [1,2,7,8]
          try:
             self.data = pickle.load(open(filename, "rb"))
          except (OSError, IOError) as e:
             for pg in page_list:
                 self.scrape_page(pg)
             self.dump_data(filename)

      def print_table(self):
          row = 0
          for row in self.data: 
              if not row : continue
              print (row, "\n")

      def start_driver(self):
          self.driver = webdriver.PhantomJS()
          self.driver.get(self.url)

      def close_driver(self):
          self.driver.quit()

      def copy_fulldata(self, dest):
          wb = Workbook()
          ws = wb.active
          #numrows = len(self.data)
          #numcols = max(len(x) for x in self.data)
          for i,row in enumerate(self.data):
              for j,col in enumerate(row):
                  ws.cell(row=i+1, column=j+1, value=col)
          wb.save(dest)

      def oper_list_of_list(self, data, name, oper):
          if oper :
             tempdata = [sum(i) for i in zip(*data) if not str in [type(e) for e in i]] 
          else:
             tempdata = [i[0] - i[1] for i in zip(*data) if not str in [type(e) for e in i]] 
          tempdata.insert(0,name)
          return tempdata
          

      def calculate_datarow(self,key,name):
          datarow = []
          if len(re.split(' \+ | - ', key)) < 2 :
             for drow in self.data:
                  if drow[0] == key:
                      datarow = drow
          else:
             tempdata = []
             delimt = ' - '
             if ' + ' in key : delimt = ' + ' 
             keys = key.split(delimt)
             for k in keys:
                for drow in self.data:
                  if drow[0] == k:
                     tempdata.append(drow)
             if delimt == ' + ':
                tempdata[0] = self.oper_list_of_list(tempdata, name, True)
             if delimt == ' - ':
                tempdata[1] = self.oper_list_of_list(tempdata[1:], name, True)
                tempdata = tempdata[:2]
                tempdata[0] = self.oper_list_of_list(tempdata, name, False) 
             datarow = tempdata[0]
          return datarow

      def copy_cellformat(self,incell, outcell):
          if incell.has_style:
             outcell.font = copy(incell.font)
             outcell.border = copy(incell.border)
             outcell.fill = copy(incell.fill)
             outcell.number_format = copy(incell.number_format)
             outcell.protection = copy(incell.protection)
             outcell.alignment = copy(incell.alignment)
              
      def update_mysheet(self,wb):
          ws = wb.active
          for row in ws.rows:
              if not isinstance(row[0].value,str):continue
              key = key_map.get(row[0].value.strip())
              if not key: continue
              datarow = self.calculate_datarow(key, row[0].value)
              for idx, datacol in enumerate(datarow):
                  if not idx: continue
                  cell = row[idx+1]
                  col = column_index_from_string(cell.column)
                  if type(datacol) != float:
                     newcell = ws.cell(row=cell.row,column=col, value=float(datacol.replace(',','')))
                  else :
                     newcell = ws.cell(row=cell.row,column=col, value=float(datacol))
                     self.copy_cellformat(cell, newcell)
      
      def zap_mysheet(self, ws):
          for row in ws.rows:
              for cell in row:
                  if isinstance(cell.value,float):
                      dcell = ws.cell(row=cell.row, column=column_index_from_string(cell.column), value=0.0)
                      self.copy_cellformat(cell, dcell)

      def copy_mysheet(self,src, dest, sheetname):
          dwb = Workbook()
          dws = dwb.active
          swb = load_workbook(filename = src, keep_vba=True)
          sws = swb.get_sheet_by_name(sheetname)
          dws.title = sws.title
          dws.sheet_view.showGridLines = False
          for row in sws.rows:
              for cell in row:
                  dcell = dws.cell(row=cell.row, column=column_index_from_string(cell.column), value=cell.value)
                  self.copy_cellformat(cell, dcell)
          self.zap_mysheet(dws)
          self.update_mysheet(dwb)
          dwb.save(dest)

      def get_list_idx(self, key, data):
          for idx, element in enumerate(data):
              if key in element: return idx

      def sort_list(self, data1,data2):
          data = data1[1:] + data2[1:]
          data = list(filter(None,data))
          data.sort(key=lambda x: x[0])
          data = data1[:1] + data2[:1] + data
          return data
              
      def merge_sublists(self, data1, data2):
          data = self.sort_list(data1,data2) 
          data_iter = enumerate(data)
          for i,row in data_iter:
               if row != data[-1] and data[i][0] == data[i+1][0]:
                  if data[i] == data[i+1]:
                     self.data.append(data[i])
                  else:
                     self.data.append(data[i] + data[i+1][1:])
                  next(data_iter)
               else:
                  if i != len(data)-1 and data[i] == data[i+1]:
                     continue
                  if len(row) == 1 : 
                     self.data.append(row)
                     continue
                  if row in data1 :
                     self.data.append(row + ['0.00'] * 5)
                  else :
                     self.data.append(row[:1] + ['0.00'] * 5 + row[1:])

      def cleanze_data(self, data):
          first = True
          data = list(filter(None,data))
          if len(data) < 5 : return None
          for row in data[:]:
              if any(re.search('12 mths',ele) is not None for ele in row):
                    data.remove(row)
                    continue
              if any(re.search('Mar \d{2}',ele) is not None for ele in row):
                 if first:
                    row.insert(0,'YEAR')
                    first = False
                 else:
                    data.remove(row)
          return data

      def change_data_type(self):
          for row in self.data:
              for idx, ele in enumerate(row):
                  #if not idx: continue
                  try:
                    if type(row[idx]) != float:
                       row[idx] = float(row[idx].replace(',','')) 
                  except ValueError: 
                    pass

      def merge_lists(self, data1, data2, num):
          if num == 1 : keyref = bs_keyref
          if num == 2 : keyref = pl_keyref
          if num == 7 : keyref = cf_keyref
          if num == 8 : keyref = ratio_keyref 

          data1 = self.cleanze_data(data1)
          data2 = self.cleanze_data(data2)

          if not data1 and data2: return 
          if not data1 : data1 = data2
          if not data2 : data2 = data1

          previdx1 = 0
          curidx1 = 0
          previdx2 = 0
          curidx2 = 0
          for ele in keyref:
              curridx1 = self.get_list_idx(ele, data1)
              curridx2 = self.get_list_idx(ele, data2)
              if not curridx1 or not curridx2 : 
                 print('%s not found' % ele)  
              self.merge_sublists(data1[previdx1:curridx1],data2[previdx2:curridx2])
              previdx1 = curridx1
              previdx2 = curridx2

              
      def pull_data(self, page_source):
          data = []
          soup = BeautifulSoup(page_source,"lxml")
          divTags = soup.find_all('div', {'class':'boxBg'})

          for tag in divTags:
               tables = tag.find_all('table',attrs={'class':'table4'})
               for table in tables:
                   table_bodys = table.find_all('tbody')
                   for tablebody in table_bodys:
                       rows = tablebody.find_all('tr')
                       for row in rows:
                           cols = row.find_all('td')
                           cols = [ele.text.strip() for ele in cols]
                           data.append([ele for ele in cols if ele]) # Get rid of empty values
          return data

      def scrape_page(self,num):
          navigXpath = ['''//*[@id="slider"]/dt[7]/a''',
                '''//*[@id="slider"]/dd[3]/ul/li['''+str(num)+''']/a''',
		'''//*[@id="mc_mainWrapper"]/div[2]/div[2]/div[3]/div[2]/div[2]/div[2]/div[1]/div/div[1]/div[1]/table/tbody/tr/td/a/b''']
          try:
             self.driver.find_element_by_xpath(navigXpath[0]).click()
             self.driver.find_element_by_xpath(navigXpath[1]).click()
             yr1 = self.pull_data(self.driver.page_source)
             self.driver.find_element_by_xpath(navigXpath[2]).click()
             yr2 = self.pull_data(self.driver.page_source)
             self.merge_lists(yr1, yr2, num)
             self.change_data_type()
          except NoSuchElementException:
             print ("Webpage Not Accessible, Try again after some time")

      def scrape_all_pages(self,pages):
          for i in range(pages):
               self.scrape_page(i+1)
          self.close_driver()
          

